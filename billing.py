#!/usr/bin/env python3
"""
billing_app_full_with_products.py
Single-file billing application with themes, PDF export, print support, product suggestions,
keyboard navigation, database storage, undo/redo, import/export, backups, product master tab, and more.

Requirements:
    pip install openpyxl reportlab PyPDF2
Optional (Windows printing):
    pip install pywin32
Optional for image preview:
    pip install pillow
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from datetime import datetime
import os
import sys
import tempfile
import json
import sqlite3
import threading
import time
import csv
import logging
import zipfile
import shutil

from openpyxl import Workbook, load_workbook

# win32print used for printer enumeration & printing on Windows (optional)
try:
    import win32print
except Exception:
    win32print = None

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from PyPDF2 import PdfReader, PdfWriter

# Optional Pillow for image thumbnail preview (not required)
try:
    from PIL import Image, ImageTk
except Exception:
    Image = None
    ImageTk = None

# ------------------- Configuration -------------------
PRODUCT_FILE = "product_master.xlsx"
BILLS_FILE = "bills.xlsx"
CUSTOMER_FILE = "customers.xlsx"
BILL_COUNTER_FILE = "bill_counter.txt"
PDF_FOLDER = "bills_pdf"
DB_FILE = "billing_app.db"
LOG_FILE = "billing_app.log"
USERS_FILE = "users.json"
SETTINGS_FILE = "settings.json"

SHOP_NAME = "P.MUTHUGANESAN NADAR TEXTILE AND READYMADE"
SHOP_ADDRESS = "103b kamachi amman sanathi street east raja veethi kanchipuram"
SHOP_PHONE = "04447791355 / 9944369227"

DEFAULT_LOW_STOCK_THRESHOLD = 100
DEFAULT_GST_PERCENT = 18.0

# ------------------- Logging -------------------
logging.basicConfig(filename=LOG_FILE, level=logging.INFO,
                    format="%(asctime)s %(levelname)s: %(message)s")

# ------------------- Helpers -------------------
def ensure_pdf_folder(folder=PDF_FOLDER):
    if not os.path.exists(folder):
        os.makedirs(folder)
    return folder

def get_next_bill_no():
    """Read bill counter file, increment, save and return (string, int)."""
    if os.path.exists(BILL_COUNTER_FILE):
        with open(BILL_COUNTER_FILE, "r") as f:
            try:
                last_no = int(f.read().strip())
            except:
                last_no = 0
    else:
        last_no = 0
    next_no = last_no + 1
    with open(BILL_COUNTER_FILE, "w") as f:
        f.write(str(next_no))
    return f"BILL-{next_no:06d}", next_no

def get_next_pdf_filename():
    ensure_pdf_folder()
    existing = [f for f in os.listdir(PDF_FOLDER) if f.endswith(".pdf")]
    if not existing:
        return os.path.join(PDF_FOLDER, "001.pdf")
    numbers = []
    for f in existing:
        name, _ = os.path.splitext(f)
        if name.isdigit():
            numbers.append(int(name))
    next_num = max(numbers) + 1 if numbers else 1
    return os.path.join(PDF_FOLDER, f"{next_num:03d}.pdf")

def read_pdf(file_path):
    if not os.path.exists(file_path):
        messagebox.showerror("Error", f"{file_path} does not exist")
        return ""
    reader = PdfReader(file_path)
    text = ""
    for page in reader.pages:
        try:
            page_text = page.extract_text()
        except Exception:
            page_text = None
        if page_text:
            text += page_text + "\n"
    return text

def merge_pdfs(pdf_list, output_file):
    writer = PdfWriter()
    for pdf_path in pdf_list:
        if not os.path.exists(pdf_path):
            continue
        reader = PdfReader(pdf_path)
        for page in reader.pages:
            writer.add_page(page)
    with open(output_file, "wb") as f:
        writer.write(f)
    messagebox.showinfo("Success", f"Merged PDFs saved as {output_file}")

# ------------------- Themes -------------------
THEMES = {
    "Light": {
        "bg": "#f7f9fc",
        "panel": "#eaf0fb",
        "accent": "#1976d2",
        "text": "#1f2937",
        "tree_even": "#ffffff",
        "tree_odd": "#f3f6fb",
        "tree_low": "#ffefef",
        "preview_bg": "#ffffff",
        "font": ("Segoe UI", 10),
        "mono": ("Consolas", 10)
    },
    "Dark": {
        "bg": "#0f1724",
        "panel": "#0b1220",
        "accent": "#1e88e5",
        "text": "#e6eef8",
        "tree_even": "#0b1220",
        "tree_odd": "#0f1626",
        "tree_low": "#3b0f0f",
        "preview_bg": "#0b1220",
        "font": ("Segoe UI", 10),
        "mono": ("Consolas", 10)
    },
    "Warm": {
        "bg": "#fcfbf8",
        "panel": "#f6efe6",
        "accent": "#ef6c00",
        "text": "#3b2f2f",
        "tree_even": "#fffaf5",
        "tree_odd": "#fff3e0",
        "tree_low": "#ffddd2",
        "preview_bg": "#fffaf0",
        "font": ("Segoe UI", 10),
        "mono": ("Consolas", 10)
    }
}
THEME_ORDER = ["Light", "Dark", "Warm"]

# ------------------- DB Schema & Utilities -------------------
DB_SCHEMA_VERSION = 1

def init_db(db_file=DB_FILE):
    """Initialize sqlite DB with tables for products, bills, and settings."""
    conn = sqlite3.connect(db_file)
    cur = conn.cursor()
    # products
    cur.execute("""
    CREATE TABLE IF NOT EXISTS products (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sku TEXT,
        name TEXT UNIQUE,
        category TEXT,
        brand TEXT,
        size TEXT,
        color TEXT,
        hsn TEXT,
        mrp REAL,
        rate REAL,
        wholesale REAL,
        super_wholesale REAL,
        discount REAL,
        qty INTEGER,
        image_path TEXT,
        notes TEXT
    )
    """)
    # categories
    cur.execute("""
    CREATE TABLE IF NOT EXISTS categories (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE
    )
    """)
    # bills
    cur.execute("""
    CREATE TABLE IF NOT EXISTS bills (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        bill_no TEXT,
        date TEXT,
        customer TEXT,
        phone TEXT,
        subtotal REAL,
        gst REAL,
        total REAL,
        paid REAL,
        due REAL
    )
    """)
    # bill_items
    cur.execute("""
    CREATE TABLE IF NOT EXISTS bill_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        bill_id INTEGER,
        product_name TEXT,
        mrp REAL,
        rate REAL,
        discount REAL,
        qty INTEGER,
        total REAL,
        FOREIGN KEY(bill_id) REFERENCES bills(id)
    )
    """)
    # settings
    cur.execute("""
    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT
    )
    """)
    # meta
    cur.execute("""
    CREATE TABLE IF NOT EXISTS meta (
        k TEXT PRIMARY KEY,
        v TEXT
    )
    """)
    # store schema version
    cur.execute("INSERT OR REPLACE INTO meta (k, v) VALUES (?, ?)", ("schema_version", str(DB_SCHEMA_VERSION)))
    conn.commit()
    conn.close()

def save_product_to_db(p, db_file=DB_FILE):
    """Insert or update a product dict into DB."""
    conn = sqlite3.connect(db_file)
    cur = conn.cursor()
    # try update by name, else insert
    cur.execute("SELECT id FROM products WHERE name=?", (p.get("name"),))
    r = cur.fetchone()
    if r:
        pid = r[0]
        cur.execute("""UPDATE products SET sku=?, category=?, brand=?, size=?, color=?, hsn=?, mrp=?, rate=?, wholesale=?, super_wholesale=?,
                       discount=?, qty=?, image_path=?, notes=? WHERE id=?""",
                    (p.get("sku"), p.get("category"), p.get("brand"), p.get("size"), p.get("color"), p.get("hsn"),
                     p.get("mrp"), p.get("rate"), p.get("wholesale"), p.get("super_wholesale"), p.get("discount"),
                     p.get("qty"), p.get("image_path"), p.get("notes"), pid))
    else:
        cur.execute("""INSERT INTO products (sku,name,category,brand,size,color,hsn,mrp,rate,wholesale,super_wholesale,discount,qty,image_path,notes)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (p.get("sku"), p.get("name"), p.get("category"), p.get("brand"), p.get("size"), p.get("color"),
                     p.get("hsn"), p.get("mrp"), p.get("rate"), p.get("wholesale"), p.get("super_wholesale"),
                     p.get("discount"), p.get("qty"), p.get("image_path"), p.get("notes")))
    conn.commit()
    conn.close()

def delete_product_from_db(name, db_file=DB_FILE):
    conn = sqlite3.connect(db_file)
    cur = conn.cursor()
    cur.execute("DELETE FROM products WHERE name=?", (name,))
    conn.commit()
    conn.close()

def load_products_from_db(db_file=DB_FILE):
    conn = sqlite3.connect(db_file)
    cur = conn.cursor()
    cur.execute("SELECT sku,name,category,brand,size,color,hsn,mrp,rate,wholesale,super_wholesale,discount,qty,image_path,notes FROM products")
    rows = cur.fetchall()
    conn.close()
    products = {}
    for r in rows:
        products[str(r[1]).lower()] = {
            "sku": r[0],
            "name": r[1],
            "category": r[2],
            "brand": r[3],
            "size": r[4],
            "color": r[5],
            "hsn": r[6],
            "mrp": r[7] or 0,
            "rate": r[8] or 0,
            "wholesale": r[9] or 0,
            "super_wholesale": r[10] or 0,
            "discount": r[11] or 0,
            "qty": r[12] or 0,
            "image_path": r[13],
            "notes": r[14]
        }
    return products

# ------------------- Undo/Redo -------------------
class UndoRedoStack:
    def __init__(self, maxlen=200):
        self.stack = []
        self.index = -1
        self.maxlen = maxlen

    def push(self, snapshot):
        if self.index < len(self.stack) - 1:
            self.stack = self.stack[:self.index+1]
        self.stack.append(json.dumps(snapshot, default=str))
        if len(self.stack) > self.maxlen:
            self.stack.pop(0)
        self.index = len(self.stack) - 1

    def can_undo(self):
        return self.index > 0

    def can_redo(self):
        return self.index < len(self.stack) - 1

    def undo(self):
        if not self.can_undo():
            return None
        self.index -= 1
        return json.loads(self.stack[self.index])

    def redo(self):
        if not self.can_redo():
            return None
        self.index += 1
        return json.loads(self.stack[self.index])

# ------------------- CSV Import / Export for products -------------------
def import_products_csv(file_path, db_file=DB_FILE):
    products = {}
    with open(file_path, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for r in reader:
            name = r.get("name") or r.get("Name") or r.get("product") or r.get("Product") or r.get("sku")
            if not name:
                keys = list(r.keys())
                if keys:
                    name = r.get(keys[0])
            if not name:
                continue
            try:
                mrp = float(r.get("mrp", 0) or 0)
                rate = float(r.get("rate", 0) or 0)
                wholesale = float(r.get("wholesale", 0) or 0)
                super_wholesale = float(r.get("super_wholesale", 0) or 0)
                discount = float(r.get("discount", 0) or 0)
                qty = int(float(r.get("qty", 0) or 0))
            except:
                mrp = rate = wholesale = super_wholesale = discount = 0.0
                qty = 0
            p = {
                "sku": r.get("sku"),
                "name": name,
                "category": r.get("category") or r.get("Category"),
                "brand": r.get("brand"),
                "size": r.get("size"),
                "color": r.get("color"),
                "hsn": r.get("hsn"),
                "mrp": mrp,
                "rate": rate,
                "wholesale": wholesale,
                "super_wholesale": super_wholesale,
                "discount": discount,
                "qty": qty,
                "image_path": r.get("image_path"),
                "notes": r.get("notes")
            }
            save_product_to_db(p, db_file=db_file)
            products[name.lower()] = p
    return products

def export_products_csv(out_file, products):
    with open(out_file, "w", newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(["name", "sku", "category", "brand", "size", "color", "hsn", "mrp", "rate", "wholesale", "super_wholesale", "discount", "qty", "image_path", "notes"])
        for name, p in products.items():
            w.writerow([p.get("name") or name, p.get("sku"), p.get("category"), p.get("brand"), p.get("size"), p.get("color"), p.get("hsn"), p.get("mrp"), p.get("rate"), p.get("wholesale"), p.get("super_wholesale"), p.get("discount"), p.get("qty"), p.get("image_path"), p.get("notes")])
    return out_file

# ------------------- Backup / Restore -------------------
def backup_project(backup_path=None):
    if not backup_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"billing_backup_{ts}.zip"
    files_to_add = []
    for fn in (PRODUCT_FILE, BILLS_FILE, CUSTOMER_FILE, BILL_COUNTER_FILE, DB_FILE, LOG_FILE, SETTINGS_FILE):
        if os.path.exists(fn):
            files_to_add.append(fn)
    if os.path.isdir(PDF_FOLDER):
        for pdf in os.listdir(PDF_FOLDER):
            files_to_add.append(os.path.join(PDF_FOLDER, pdf))
    with zipfile.ZipFile(backup_path, "w", zipfile.ZIP_DEFLATED) as z:
        for f in files_to_add:
            try:
                z.write(f, os.path.basename(f))
            except Exception:
                logging.exception(f"Failed to add {f} to backup")
    return backup_path

def restore_project_from_zip(zip_path):
    if not os.path.exists(zip_path):
        raise FileNotFoundError(zip_path)
    with zipfile.ZipFile(zip_path, "r") as z:
        z.extractall(".")
    return True

# ------------------- Settings -------------------
DEFAULT_SETTINGS = {
    "auto_backup": False,
    "auto_backup_interval_minutes": 60,
    "recent_files": [],
    "last_theme": "Light",
    "page_size": 200
}

def load_settings():
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass
    # try DB fallback
    try:
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT value FROM settings WHERE key=?", ("app_settings",))
        r = cur.fetchone()
        conn.close()
        if r:
            return json.loads(r[0])
    except:
        pass
    return DEFAULT_SETTINGS.copy()

def save_settings(settings):
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(settings, f, indent=2)
    # save to DB as well
    try:
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", ("app_settings", json.dumps(settings)))
        conn.commit()
        conn.close()
    except:
        pass

# ------------------- Progress UI (for heavy imports) -------------------
class ProgressWindow:
    def __init__(self, parent, title="Working...", maxval=100):
        self.win = tk.Toplevel(parent)
        self.win.title(title)
        self.win.geometry("400x80")
        self.win.transient(parent)
        self.win.grab_set()
        self.pb = ttk.Progressbar(self.win, orient="horizontal", length=360, mode="determinate", maximum=maxval)
        self.pb.pack(pady=16, padx=16)
        self.label = tk.Label(self.win, text="")
        self.label.pack()

    def set(self, val, text=None):
        try:
            self.pb['value'] = val
            if text:
                self.label.config(text=text)
            self.win.update_idletasks()
        except:
            pass

    def close(self):
        try:
            self.win.destroy()
        except:
            pass

# ------------------- Main BillingApp -------------------
class BillingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Billing Software")
        try:
            self.root.state("zoomed")
        except:
            pass

        # Core data
        self.sub_total = 0.0
        self.gst_percent = tk.DoubleVar(value=DEFAULT_GST_PERCENT)
        self.total = 0.0
        self.paid_amount = tk.DoubleVar(value=0.0)
        self.due_amount = tk.DoubleVar(value=0.0)
        self.low_stock_threshold = DEFAULT_LOW_STOCK_THRESHOLD
        # products loaded from DB or xlsx
        try:
            init_db()
        except:
            pass
        try:
            self.products = load_products_from_db()
        except:
            self.products = {}
        self.items_in_bill = []

        # Theme
        self.current_theme_name = "Light"
        self.current_theme = THEMES[self.current_theme_name]

        # Style
        self.style = ttk.Style()
        try:
            self.style.theme_use("clam")
        except:
            pass

        # Build UI
        self.build_ui()

        # Shortcuts
        self.root.bind("<Delete>", lambda e: self.delete_item())
        self.root.bind("<Control-d>", lambda e: self.duplicate_selected_item())
        self.root.bind("<Control-e>", lambda e: self.edit_selected_item())
        self.root.bind("<Return>", lambda e: self.enter_key_pressed(e))
        self.root.bind("<Control-n>", lambda e: self.new_client())
        self.root.bind("<Control-Shift-P>", lambda e: self.open_products_tab())  # Ctrl+Shift+P for products
        self.root.bind("<Control-p>", lambda e: self.show_print_tab())
        self.root.bind("<Control-s>", lambda e: self.print_and_save_pdf())
        self.root.bind("<Control-S>", lambda e: self.save_customer())  # Ctrl+Shift+S
        self.root.bind("<Control-Up>", lambda e: self.move_item_up())
        self.root.bind("<Control-Down>", lambda e: self.move_item_down())
        self.root.bind("<F5>", lambda e: self.refresh_printer_list())
        self.root.bind("<F9>", lambda e: self.reload_products())
        self.root.bind("<Alt-t>", lambda e: self.cycle_theme())
        self.root.bind("<Control-l>", lambda e: self.threshold_entry.focus_set())
        self.root.bind("<Control-i>", lambda e: self.item_name.focus_set())
        self.root.bind("<Control-b>", lambda e: self.bill_preview.focus_set())
        self.root.bind("<Control-r>", lambda e: self.update_totals())
        self.root.bind("<Control-m>", lambda e: self.open_merge_pdf())
        self.root.bind("<Control-h>", lambda e: self.show_shortcuts_help())

        # Undo/Redo
        self._ur_stack = UndoRedoStack(maxlen=200)
        self._push_snapshot = lambda: self._ur_stack.push(self._snapshot())
        self._push_snapshot()

        # Auto-save stub
        self.auto_save_interval = 5000
        self.root.after(self.auto_save_interval, self.auto_save_bill)

    # ------- Snapshot for undo/redo -------
    def _snapshot(self):
        return {
            "items": list(self.items_in_bill),
            "cust_name": self.cust_name.get() if hasattr(self, "cust_name") else "",
            "cust_phone": self.cust_phone.get() if hasattr(self, "cust_phone") else "",
            "paid": self.paid_amount.get(),
            "gst": self.gst_percent.get()
        }

    # ---------------- UI build ----------------
    def build_ui(self):
        self.root.configure(bg=self.current_theme["bg"])

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)

        # Main & Print frames & Products frame
        self.main_frame = tk.Frame(self.notebook, bg=self.current_theme["bg"])
        self.print_frame = tk.Frame(self.notebook, bg=self.current_theme["bg"])
        self.products_frame = tk.Frame(self.notebook, bg=self.current_theme["bg"])
        self.notebook.add(self.main_frame, text="🧾 Billing")
        self.notebook.add(self.print_frame, text="🖨️ Print Bill")
        self.notebook.add(self.products_frame, text="📦 Products Master")

        # Main frame sections
        self.build_customer_frame()
        self.build_item_frame()
        self.build_toolbar()
        self.build_treeview()
        self.build_bottom_frame()
        self.build_extra_tool_row()

        # Print frame
        self.build_print_preview_with_printer_select()

        # Products Master tab
        self.build_products_tab()

        # Apply theme
        self.apply_theme(self.current_theme_name)

    # ---------------- Billing UI pieces ----------------
    def build_customer_frame(self):
        cust_frame = tk.LabelFrame(self.main_frame, text="Customer Details", padx=10, pady=10)
        cust_frame.pack(fill="x", padx=10, pady=8)
        tk.Label(cust_frame, text="Name:").grid(row=0, column=0, sticky="w")
        self.cust_name = tk.Entry(cust_frame, width=30)
        self.cust_name.grid(row=0, column=1, padx=6)
        self.cust_name.insert(0, "Customer")
        tk.Label(cust_frame, text="Phone:").grid(row=0, column=2, sticky="w")
        self.cust_phone = tk.Entry(cust_frame, width=20)
        self.cust_phone.grid(row=0, column=3, padx=6)
        ttk.Button(cust_frame, text="Save Customer", command=self.save_customer).grid(row=0, column=4, padx=12)

    def build_item_frame(self):
        item_frame = tk.LabelFrame(self.main_frame, text="Add Item", padx=10, pady=10)
        item_frame.pack(fill="x", padx=10, pady=6)
        tk.Label(item_frame, text="Item:").grid(row=0, column=0, sticky="w")
        self.item_name = tk.Entry(item_frame, width=40)
        self.item_name.grid(row=0, column=1, padx=6)
        self.item_name.bind("<KeyRelease>", self.show_suggestions)
        self.item_name.bind("<Down>", self.move_down)
        self.item_name.bind("<Up>", self.move_up)
        self.item_name.bind("<Escape>", lambda e: self.suggestion_box.grid_remove())
        self.item_name.bind("<Return>", self.enter_in_item_name)

        self.suggestion_box = tk.Listbox(item_frame, height=6)
        self.suggestion_box.grid(row=1, column=1, sticky="w", padx=6)
        self.suggestion_box.bind("<<ListboxSelect>>", self.fill_from_suggestion)
        self.suggestion_box.bind("<Return>", self.fill_from_suggestion)
        self.suggestion_box.grid_remove()

        tk.Label(item_frame, text="MRP:").grid(row=0, column=2, sticky="w")
        self.item_mrp = tk.Entry(item_frame, width=10)
        self.item_mrp.grid(row=0, column=3, padx=6)
        tk.Label(item_frame, text="Rate:").grid(row=0, column=4, sticky="w")
        self.item_rate = tk.Entry(item_frame, width=10)
        self.item_rate.grid(row=0, column=5, padx=6)
        tk.Label(item_frame, text="Disc%:").grid(row=0, column=6, sticky="w")
        self.item_discount = tk.Entry(item_frame, width=8)
        self.item_discount.grid(row=0, column=7, padx=6)
        tk.Label(item_frame, text="Qty:").grid(row=0, column=8, sticky="w")
        self.item_qty = tk.Entry(item_frame, width=6)
        self.item_qty.grid(row=0, column=9, padx=6)
        self.item_qty.bind("<Return>", lambda e: self.add_item())
        ttk.Button(item_frame, text="Add Item", command=self.add_item).grid(row=0, column=10, padx=12)

    def build_toolbar(self):
        toolbar = tk.Frame(self.main_frame, pady=6)
        toolbar.pack(fill="x", padx=10)
        ttk.Button(toolbar, text="Add (Enter)", command=self.add_item).pack(side="left", padx=4)
        ttk.Button(toolbar, text="Delete (Del)", command=self.delete_item).pack(side="left", padx=4)
        ttk.Button(toolbar, text="New Client (Ctrl+N)", command=self.new_client).pack(side="left", padx=4)
        ttk.Button(toolbar, text="Print (Ctrl+P)", command=self.show_print_tab).pack(side="left", padx=4)
        # Product Manager quick button
        ttk.Button(toolbar, text="Products (Ctrl+Shift+P)", command=self.open_products_tab).pack(side="left", padx=4)
        ttk.Button(toolbar, text="Remove Low Stock", command=self.delete_low_stock_items).pack(side="left", padx=4)
        ttk.Button(toolbar, text="Read PDF", command=self.open_read_pdf).pack(side="left", padx=4)
        ttk.Button(toolbar, text="Merge PDFs", command=self.open_merge_pdf).pack(side="left", padx=4)

        right_frame = tk.Frame(toolbar)
        right_frame.pack(side="right")
        tk.Label(right_frame, text="LowStockThreshold:").pack(side="left", padx=6)
        self.threshold_entry = tk.Entry(right_frame, width=6)
        self.threshold_entry.pack(side="left")
        self.threshold_entry.insert(0, str(self.low_stock_threshold))
        ttk.Button(right_frame, text="Set Threshold", command=self.set_low_stock_threshold).pack(side="left", padx=6)
        tk.Label(right_frame, text="GST %:").pack(side="left", padx=6)
        tk.Entry(right_frame, textvariable=self.gst_percent, width=6).pack(side="left")
        tk.Label(right_frame, text="Theme:").pack(side="left", padx=(12,4))
        self.theme_var = tk.StringVar(value=self.current_theme_name)
        ttk.OptionMenu(right_frame, self.theme_var, self.current_theme_name, *THEMES.keys(), command=self.on_theme_select).pack(side="left")
        ttk.Button(right_frame, text="Cycle Theme (Alt+T)", command=self.cycle_theme).pack(side="left", padx=6)
        ttk.Button(right_frame, text="Shortcuts (Ctrl+H)", command=self.show_shortcuts_help).pack(side="left", padx=6)

    def build_extra_tool_row(self):
        tool_frame = tk.Frame(self.main_frame, bg=self.current_theme.get("bg"))
        tool_frame.pack(fill="x", padx=10, pady=4)
        ttk.Button(tool_frame, text="Undo (Ctrl+Z)", command=lambda: getattr(self, "undo", lambda: None)()).pack(side="left", padx=4)
        ttk.Button(tool_frame, text="Redo (Ctrl+Y)", command=lambda: getattr(self, "redo", lambda: None)()).pack(side="left", padx=4)
        ttk.Button(tool_frame, text="Import Products CSV", command=self.import_products_ui).pack(side="left", padx=4)
        ttk.Button(tool_frame, text="Export Products CSV", command=self.export_products_ui).pack(side="left", padx=4)
        ttk.Button(tool_frame, text="Backup Now", command=lambda: messagebox.showinfo("Backup", f"Backup created: {backup_project()}")).pack(side="left", padx=4)
        ttk.Button(tool_frame, text="Preferences", command=lambda: open_preferences_dialog(self.root, load_settings())).pack(side="left", padx=4)

    def build_treeview(self):
        columns = ("Item","MRP","Rate","Discount","Qty","Total")
        self.tree = ttk.Treeview(self.main_frame, columns=columns, show="headings", height=14)
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=300 if col=="Item" else 90, anchor="center")
        self.tree.pack(fill="both", expand=True, padx=10, pady=10)
        self.tree.tag_configure("evenrow")
        self.tree.tag_configure("oddrow")
        self.tree.tag_configure("lowstock")
        self.tree.bind("<Double-1>", lambda e: self.edit_selected_item())

    def build_bottom_frame(self):
        bottom_frame = tk.Frame(self.main_frame)
        bottom_frame.pack(fill="x", padx=10, pady=(0,10))
        ttk.Button(bottom_frame, text="Clear All (Ctrl+C)", command=self.clear_all).pack(side="left")
        self.sub_label = tk.Label(bottom_frame, text="SubTotal: 0.00", font=self.current_theme.get("font"))
        self.sub_label.pack(side="right", padx=10)
        self.gst_label = tk.Label(bottom_frame, text="GST: 0.00", font=self.current_theme.get("font"))
        self.gst_label.pack(side="right", padx=10)
        self.total_label = tk.Label(bottom_frame, text="Total: 0.00", font=self.current_theme.get("font"))
        self.total_label.pack(side="right", padx=10)
        self.item_count_label = tk.Label(bottom_frame, text="Items: 0", font=self.current_theme.get("font"))
        self.item_count_label.pack(side="right", padx=10)
        tk.Label(bottom_frame, text="Paid:").pack(side="right", padx=6)
        paid_entry = tk.Entry(bottom_frame, textvariable=self.paid_amount, width=10)
        paid_entry.pack(side="right")
        paid_entry.bind("<KeyRelease>", lambda e: self.update_due())
        tk.Label(bottom_frame, text="Due:").pack(side="right", padx=6)
        self.due_label = tk.Label(bottom_frame, text="0.00", font=self.current_theme.get("font"))
        self.due_label.pack(side="right", padx=6)

    def build_print_preview_with_printer_select(self):
        self.bill_preview = tk.Text(self.print_frame, font=self.current_theme.get("mono"), width=90, height=25)
        self.bill_preview.pack(fill="both", expand=True, padx=10, pady=(10,6))
        printer_frame = tk.Frame(self.print_frame)
        printer_frame.pack(fill="x", padx=10, pady=(0,6))
        tk.Label(printer_frame, text="Select Printer:").pack(side="left", padx=(0,6))
        self.printer_var = tk.StringVar()
        self.printer_combo = ttk.Combobox(printer_frame, textvariable=self.printer_var, state="readonly", width=60)
        self.printer_combo.pack(side="left", padx=(0,6))
        ttk.Button(printer_frame, text="Refresh Printers (F5)", command=self.refresh_printer_list).pack(side="left", padx=(6,0))
        btn_frame = tk.Frame(self.print_frame)
        btn_frame.pack(pady=(0,10))
        ttk.Button(btn_frame, text="Print", command=self.print_bill_to_printer).pack(side="left", padx=6)
        ttk.Button(btn_frame, text="Save PDF (Ctrl+S)", command=self.print_and_save_pdf).pack(side="left", padx=6)
        ttk.Button(btn_frame, text="Cancel", command=lambda: self.notebook.select(self.main_frame)).pack(side="left", padx=6)

    # ---------------- Theme handlers ----------------
    def apply_theme(self, theme_name):
        if theme_name not in THEMES:
            return
        self.current_theme_name = theme_name
        self.current_theme = THEMES[theme_name]
        t = self.current_theme
        try:
            self.root.configure(bg=t["bg"])
        except:
            pass
        try:
            self.style.configure("TNotebook", background=t["bg"])
            self.style.configure("TNotebook.Tab", font=t["font"], padding=[10,6])
            self.style.configure("TFrame", background=t["bg"])
            self.style.configure("TLabel", background=t["panel"], font=t["font"], foreground=t["text"])
            self.style.configure("TButton", font=t["font"])
        except:
            pass
        for widget in (self.main_frame, self.print_frame, self.products_frame):
            try:
                widget.configure(bg=t["bg"])
            except:
                pass
        for child in self.main_frame.winfo_children():
            try:
                child.configure(bg=t["panel"])
            except:
                pass
        try:
            self.suggestion_box.configure(bg=t["tree_even"], fg=t["text"], font=t["font"])
        except:
            pass
        try:
            self.style.configure("Treeview", font=t["font"], foreground=t["text"])
            self.tree.tag_configure("evenrow", background=t["tree_even"])
            self.tree.tag_configure("oddrow", background=t["tree_odd"])
            self.tree.tag_configure("lowstock", background=t["tree_low"])
        except:
            pass
        for lbl in (self.sub_label, self.gst_label, self.total_label, self.item_count_label, self.due_label):
            try:
                lbl.configure(bg=t["bg"], fg=t["text"], font=t["font"])
            except:
                pass
        try:
            self.bill_preview.configure(bg=t["preview_bg"], fg=t["text"], font=t["mono"])
        except:
            pass
        entries = [getattr(self,attr) for attr in ("cust_name","cust_phone","item_name","item_mrp","item_rate","item_discount","item_qty","threshold_entry") if hasattr(self,attr)]
        for e in entries:
            try:
                bg_color = "white" if theme_name == "Light" else t["panel"]
                e.configure(bg=bg_color, fg=t["text"], font=t["font"])
            except:
                pass
        try:
            self.theme_var.set(self.current_theme_name)
        except:
            pass
        self.refresh_tree()
        # also adjust product tab tree if exists
        try:
            self.refresh_products_view()
        except:
            pass

    def cycle_theme(self):
        idx = THEME_ORDER.index(self.current_theme_name)
        next_idx = (idx + 1) % len(THEME_ORDER)
        self.apply_theme(THEME_ORDER[next_idx])

    def on_theme_select(self, selected):
        self.apply_theme(selected)

    # ---------------- Printers ----------------
    def refresh_printer_list(self):
        printers = []
        default_printer = None
        if win32print:
            try:
                raw = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)
                for p in raw:
                    try:
                        pname = p[2]
                    except:
                        pname = str(p)
                    printers.append(pname)
                try:
                    default_printer = win32print.GetDefaultPrinter()
                except:
                    default_printer = None
            except:
                printers = []
                default_printer = None
        if not printers:
            if default_printer:
                printers = [default_printer]
            else:
                printers = ["<No printers found>"]
        try:
            self.printer_combo["values"] = printers
            if default_printer and default_printer in printers:
                self.printer_var.set(default_printer)
            else:
                self.printer_var.set(printers[0])
        except:
            pass

    # ---------------- Products / Suggestions ----------------
    def reload_products(self):
        try:
            p = load_products_from_db()
            if p:
                self.products = p
            else:
                # fallback to excel
                if os.path.exists(PRODUCT_FILE):
                    wb = load_workbook(PRODUCT_FILE)
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:
                            self.products[str(row[0]).lower()] = {"name":row[0],"mrp":row[1] or 0,"rate":row[2] or 0,"discount":row[3] or 0,"qty":row[4] if len(row)>4 else 0}
        except Exception:
            logging.exception("Failed to reload products")
        messagebox.showinfo("Products Reloaded", f"Loaded {len(self.products)} products")
        self.refresh_products_view()

    def show_suggestions(self, event):
        val = self.item_name.get().lower()
        if not val:
            self.suggestion_box.grid_remove()
            return
        matches = [p for p in self.products if val in p]
        if matches:
            self.suggestion_box.delete(0, tk.END)
            for m in matches[:20]:
                self.suggestion_box.insert(tk.END, m)
            self.suggestion_box.grid()
        else:
            self.suggestion_box.grid_remove()

    def fill_from_suggestion(self, event=None):
        sel = self.suggestion_box.curselection()
        if sel:
            val = self.suggestion_box.get(sel)
            prod = self.products.get(val.lower(), {})
            self.item_name.delete(0, tk.END); self.item_name.insert(0, prod.get("name", val))
            self.item_mrp.delete(0, tk.END); self.item_mrp.insert(0, prod.get("mrp", ""))
            self.item_rate.delete(0, tk.END); self.item_rate.insert(0, prod.get("rate", ""))
            self.item_discount.delete(0, tk.END); self.item_discount.insert(0, prod.get("discount", ""))
            self.suggestion_box.grid_remove()
            self.item_qty.focus()

    def move_down(self, event):
        if self.suggestion_box.size() > 0:
            self.suggestion_box.focus()
            self.suggestion_box.selection_set(0)
            self.suggestion_box.activate(0)

    def move_up(self, event):
        if self.suggestion_box.size() > 0:
            self.suggestion_box.focus()
            self.suggestion_box.selection_set(self.suggestion_box.size()-1)
            self.suggestion_box.activate(self.suggestion_box.size()-1)

    def enter_in_item_name(self, event):
        if self.suggestion_box.winfo_ismapped():
            self.fill_from_suggestion()
        else:
            self.add_item()

    # ---------------- Billing actions ----------------
    def add_item(self):
        try:
            name = self.item_name.get().strip()
            mrp = float(self.item_mrp.get() or 0)
            rate = float(self.item_rate.get() or 0)
            disc = float(self.item_discount.get() or 0)
            qty = int(self.item_qty.get() or 0)
        except Exception:
            messagebox.showerror("Error","Invalid item details")
            return
        if not name:
            messagebox.showerror("Error","Item name required")
            return
        total = round(rate * qty * (1 - disc/100), 2)
        self.items_in_bill.append({"name":name,"mrp":mrp,"rate":rate,"discount":disc,"qty":qty,"total":total})
        try: self._push_snapshot()
        except: pass
        self.update_totals()
        self.refresh_tree()
        self.item_name.delete(0, tk.END)
        self.item_mrp.delete(0, tk.END)
        self.item_rate.delete(0, tk.END)
        self.item_discount.delete(0, tk.END)
        self.item_qty.delete(0, tk.END)

    def delete_item(self):
        sel = self.tree.selection()
        if not sel:
            return
        try:
            index = int(sel[0])
        except:
            for s in sel:
                try:
                    self.tree.delete(s)
                except:
                    pass
            return
        if 0 <= index < len(self.items_in_bill):
            del self.items_in_bill[index]
        try: self._push_snapshot()
        except: pass
        self.update_totals()
        self.refresh_tree()

    def duplicate_selected_item(self):
        sel = self.tree.selection()
        if not sel:
            return
        try:
            idx = int(sel[0])
        except:
            return
        if 0 <= idx < len(self.items_in_bill):
            item = self.items_in_bill[idx].copy()
            self.items_in_bill.insert(idx+1, item)
            try: self._push_snapshot()
            except: pass
            self.refresh_tree()
            self.update_totals()

    def edit_selected_item(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Edit Item", "Select an item in the list and press Ctrl+E or double-click.")
            return
        try:
            idx = int(sel[0])
        except:
            return
        if not (0 <= idx < len(self.items_in_bill)):
            return
        item = self.items_in_bill[idx]

        new_name = simpledialog.askstring("Edit Item", "Item name:", initialvalue=item["name"], parent=self.root)
        if new_name is None:
            return
        try:
            new_mrp = float(simpledialog.askstring("Edit Item", "MRP:", initialvalue=str(item["mrp"]), parent=self.root) or item["mrp"])
            new_rate = float(simpledialog.askstring("Edit Item", "Rate:", initialvalue=str(item["rate"]), parent=self.root) or item["rate"])
            new_disc = float(simpledialog.askstring("Edit Item", "Discount%:", initialvalue=str(item["discount"]), parent=self.root) or item["discount"])
            new_qty = int(simpledialog.askstring("Edit Item", "Qty:", initialvalue=str(item["qty"]), parent=self.root) or item["qty"])
        except Exception:
            messagebox.showerror("Error", "Invalid values entered")
            return

        item.update({"name": new_name, "mrp": new_mrp, "rate": new_rate, "discount": new_disc, "qty": new_qty})
        item["total"] = round(item["rate"] * item["qty"] * (1 - item["discount"]/100), 2)
        try: self._push_snapshot()
        except: pass
        self.refresh_tree()
        self.update_totals()

    def move_item_up(self):
        sel = self.tree.selection()
        if not sel:
            return
        try:
            idx = int(sel[0])
        except:
            return
        if idx <= 0:
            return
        self.items_in_bill[idx-1], self.items_in_bill[idx] = self.items_in_bill[idx], self.items_in_bill[idx-1]
        try: self._push_snapshot()
        except: pass
        self.refresh_tree()
        self.tree.selection_set(str(idx-1))

    def move_item_down(self):
        sel = self.tree.selection()
        if not sel:
            return
        try:
            idx = int(sel[0])
        except:
            return
        if idx >= len(self.items_in_bill)-1:
            return
        self.items_in_bill[idx+1], self.items_in_bill[idx] = self.items_in_bill[idx], self.items_in_bill[idx+1]
        try: self._push_snapshot()
        except: pass
        self.refresh_tree()
        self.tree.selection_set(str(idx+1))

    def clear_all(self):
        self.items_in_bill.clear()
        try: self._push_snapshot()
        except: pass
        self.update_totals()
        self.refresh_tree()
        try:
            self.cust_name.delete(0, tk.END)
            self.cust_name.insert(0, "Customer")
            self.cust_phone.delete(0, tk.END)
        except:
            pass

    def new_client(self):
        self.clear_all()

    def update_totals(self):
        self.sub_total = sum(i["total"] for i in self.items_in_bill)
        gst_amt = round(self.sub_total * self.gst_percent.get() / 100, 2)
        self.total = self.sub_total + gst_amt
        self.sub_label.config(text=f"SubTotal: {self.sub_total:.2f}")
        self.gst_label.config(text=f"GST: {gst_amt:.2f}")
        self.total_label.config(text=f"Total: {self.total:.2f}")
        self.item_count_label.config(text=f"Items: {len(self.items_in_bill)}")
        self.update_due()

    def update_due(self):
        paid = self.paid_amount.get()
        due = self.total - paid
        self.due_label.config(text=f"{due:.2f}")

    def refresh_tree(self):
        self.tree.delete(*self.tree.get_children())
        for idx, item in enumerate(self.items_in_bill):
            tag = "lowstock" if item["qty"] <= self.low_stock_threshold else ("evenrow" if idx%2==0 else "oddrow")
            self.tree.insert("", tk.END, iid=str(idx), values=(item["name"], item["mrp"], item["rate"], item["discount"], item["qty"], item["total"]), tags=(tag,))

    # ---------------- Customer persistence ----------------
    def save_customer(self):
        name = self.cust_name.get().strip()
        phone = self.cust_phone.get().strip()
        if not name or not phone:
            messagebox.showwarning("Save Customer", "Name and phone required.")
            return
        wb = Workbook()
        if os.path.exists(CUSTOMER_FILE):
            wb = load_workbook(CUSTOMER_FILE)
        ws = wb.active
        ws.append([name, phone, str(datetime.now())])
        wb.save(CUSTOMER_FILE)
        messagebox.showinfo("Saved", "Customer saved")

    def delete_low_stock_items(self):
        self.items_in_bill = [i for i in self.items_in_bill if i["qty"] > self.low_stock_threshold]
        try: self._push_snapshot()
        except: pass
        self.update_totals()
        self.refresh_tree()

    def set_low_stock_threshold(self):
        try:
            self.low_stock_threshold = int(self.threshold_entry.get())
            messagebox.showinfo("Threshold","Low stock threshold updated")
            self.refresh_tree()
        except:
            messagebox.showerror("Error","Invalid threshold")

    # ---------------- PDF saving ----------------
    def print_and_save_pdf(self):
        bill_no, _ = get_next_bill_no()
        pdf_file = get_next_pdf_filename()
        try:
            c = canvas.Canvas(pdf_file, pagesize=A4)
            c.setFont("Helvetica-Bold", 16)
            c.drawString(60, 800, SHOP_NAME)
            c.setFont("Helvetica", 10)
            c.drawString(60, 785, SHOP_ADDRESS)
            c.drawString(60, 770, SHOP_PHONE)
            c.setFont("Helvetica", 10)
            c.drawString(400, 740, f"Bill No: {bill_no}")
            c.drawString(400, 725, f"Date: {datetime.now().strftime('%d-%m-%Y %H:%M')}")
            data = [["Item","MRP","Rate","Disc%","Qty","Total"]]
            for it in self.items_in_bill:
                data.append([it["name"], it["mrp"], it["rate"], it["discount"], it["qty"], it["total"]])
            table = Table(data, colWidths=[200,50,50,50,40,60])
            table.setStyle(TableStyle([("GRID",(0,0),(-1,-1),0.5,colors.black),
                                       ("FONT",(0,0),(-1,0),"Helvetica-Bold")] ))
            table.wrapOn(c, 40, 600)
            table.drawOn(c, 40, 600 - len(data)*18)
            y = 600 - len(data)*18 - 20
            c.drawString(40, y, f"SubTotal: {self.sub_total:.2f}")
            y -= 14
            gst_amt = self.sub_total * self.gst_percent.get() / 100
            c.drawString(40, y, f"GST: {gst_amt:.2f}")
            y -= 14
            c.drawString(40, y, f"Total: {self.total:.2f}")
            c.save()
            messagebox.showinfo("PDF Saved", f"Bill saved as {pdf_file}")
            # save to DB as well
            try:
                save_bill_to_db(self)
            except Exception:
                logging.exception("Failed to save bill to DB")
        except Exception as e:
            messagebox.showerror("PDF Error", f"Failed to create PDF: {e}")
            logging.exception("PDF generation failed")

    def open_read_pdf(self):
        file = filedialog.askopenfilename(filetypes=[("PDF Files","*.pdf")])
        if file:
            text = read_pdf(file)
            win = tk.Toplevel(self.root)
            win.title("PDF Content")
            txt = tk.Text(win, width=100, height=30)
            txt.pack(fill="both", expand=True)
            txt.insert(tk.END, text)

    def open_merge_pdf(self):
        files = filedialog.askopenfilenames(filetypes=[("PDF Files","*.pdf")])
        if files:
            out = filedialog.asksaveasfilename(defaultextension=".pdf")
            if out:
                merge_pdfs(files, out)

    # ---------------- Print preview & printing (TEXT receipt) ----------------
    def show_print_tab(self):
        self.refresh_printer_list()
        self.refresh_print_bill()
        self.notebook.select(self.print_frame)

    def refresh_print_bill(self):
        self.bill_preview.delete("1.0", tk.END)
        bill_no, _ = get_next_bill_no()
        lines = []
        lines.append(f"{SHOP_NAME}\n")
        lines.append(f"{SHOP_ADDRESS}\n")
        lines.append(f"{SHOP_PHONE}\n")
        lines.append("="*60 + "\n")
        lines.append(f"Bill No: {bill_no:<10} Date: {datetime.now().strftime('%d-%m-%Y %H:%M')}\n")
        lines.append("="*60 + "\n")
        lines.append(f"{'Item':20}{'MRP':>6}{'Rate':>7}{'Disc%':>7}{'Qty':>6}{'Total':>9}\n")
        lines.append("-"*60 + "\n")
        for i in self.items_in_bill:
            lines.append(f"{i['name'][:20]:20}{i['mrp']:>6.2f}{i['rate']:>7.2f}{i['discount']:>7.2f}{i['qty']:>6}{i['total']:>9.2f}\n")
        lines.append("-"*60 + "\n")
        lines.append(f"SubTotal: {self.sub_total:.2f}\n")
        gst_amt = self.sub_total * self.gst_percent.get() / 100
        lines.append(f"GST: {gst_amt:.2f}\n")
        lines.append(f"Total: {self.total:.2f}\n")
        paid = self.paid_amount.get()
        due = self.total - paid
        lines.append(f"Paid: {paid:.2f}  Due: {due:.2f}\n")
        lines.append("="*60 + "\n")
        self.bill_preview.insert(tk.END, "".join(lines))

    def print_bill_to_printer(self):
        text = self.bill_preview.get("1.0", tk.END)
        if not text.strip():
            messagebox.showerror("Error", "Nothing to print")
            return
        try:
            tmp_pdf = tempfile.mktemp(suffix=".pdf")
            c = canvas.Canvas(tmp_pdf, pagesize=A4)
            width, height = A4
            margin = 40
            y = height - margin
            c.setFont("Helvetica-Bold", 14)
            c.drawString(margin, y, SHOP_NAME)
            y -= 18
            c.setFont("Helvetica", 10)
            c.drawString(margin, y, SHOP_ADDRESS)
            y -= 14
            c.drawString(margin, y, SHOP_PHONE)
            y -= 20
            c.setFont("Helvetica", 9)
            c.drawString(margin, y, "-" * 95)
            y -= 16
            lines = text.splitlines()
            c.setFont("Courier", 9)
            line_height = 12
            max_lines = int((y - margin) / line_height) or 40
            idx = 0
            while idx < len(lines):
                for _ in range(max_lines):
                    if idx >= len(lines):
                        break
                    line = lines[idx]
                    while len(line) > 0:
                        piece = line[:110]
                        c.drawString(margin, y, piece)
                        y -= line_height
                        line = line[110:]
                        if y < margin + line_height:
                            break
                    if y < margin + line_height:
                        break
                    idx += 1
                if idx < len(lines):
                    c.showPage()
                    y = height - margin
                    c.setFont("Courier", 9)
            c.save()
        except Exception as e:
            messagebox.showerror("PDF Error", f"Failed to create PDF for printing: {e}")
            logging.exception("Failed to create temp PDF for printing")
            return
        try:
            selected = self.printer_var.get()
            if selected == "<No printers found>":
                selected = None
        except:
            selected = None
        if win32print and selected:
            try:
                win32print.ShellExecute(0, "print", tmp_pdf, f'/d:"{selected}"', ".", 0)
                messagebox.showinfo("Printed", f"Bill sent to printer: {selected}")
                return
            except Exception as e:
                messagebox.showwarning("Print Error", f"Could not print to '{selected}': {e}\nTrying default printer...")
                logging.exception("Targeted print failed")
        if win32print:
            try:
                default = win32print.GetDefaultPrinter()
                win32print.ShellExecute(0, "print", tmp_pdf, f'/d:"{default}"', ".", 0)
                messagebox.showinfo("Printed", f"Bill sent to default printer: {default}")
                return
            except Exception:
                try:
                    os.startfile(tmp_pdf)
                    messagebox.showinfo("PDF Ready", f"PDF opened for printing: {tmp_pdf}")
                except Exception:
                    messagebox.showinfo("PDF Saved", f"PDF saved to: {tmp_pdf}")
                return
        try:
            if sys.platform == "darwin":
                os.system(f'open "{tmp_pdf}"')
            elif os.name == "nt":
                os.startfile(tmp_pdf)
            else:
                os.system(f'xdg-open "{tmp_pdf}"')
            messagebox.showinfo("PDF Ready", f"PDF saved and opened for manual printing:\n{tmp_pdf}")
        except Exception:
            messagebox.showinfo("PDF Saved", f"PDF saved to: {tmp_pdf}\nPlease open and print manually.")

    # ---------------- Auto-save stub ----------------
    def auto_save_bill(self):
        self.root.after(self.auto_save_interval, self.auto_save_bill)

    # ---------------- Utility/key helpers ----------------
    def enter_key_pressed(self, event):
        w = self.root.focus_get()
        if w == self.tree:
            self.edit_selected_item()
        elif w in (self.item_name, self.item_mrp, self.item_rate, self.item_discount, self.item_qty):
            self.add_item()

    def show_shortcuts_help(self):
        shortcuts = [
            "Ctrl+N — New client / clear invoice",
            "Ctrl+P — Open Print tab & refresh preview",
            "Ctrl+Shift+P — Open Products Master",
            "Ctrl+S — Save bill as PDF",
            "Ctrl+Shift+S — Save customer",
            "Delete — Delete selected item",
            "Ctrl+D — Duplicate selected item",
            "Ctrl+E or Double-click — Edit selected item",
            "Ctrl+Up / Ctrl+Down — Move selected item up / down",
            "F5 — Refresh printer list",
            "F9 — Reload product master",
            "Alt+T — Cycle theme",
            "Ctrl+L — Focus LowStockThreshold entry",
            "Ctrl+I — Focus Item name (start adding)",
            "Ctrl+B — Focus Bill preview (print tab)",
            "Ctrl+R — Refresh totals",
            "Ctrl+M — Merge PDFs (opens file dialog)",
            "Ctrl+H — Show this help"
        ]
        messagebox.showinfo("Keyboard Shortcuts", "\n".join(shortcuts))

    def import_products_ui(self):
        fp = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")])
        if not fp:
            return
        try:
            # show progress window while importing
            pw = ProgressWindow(self.root, title="Importing products...", maxval=100)
            def task():
                products = import_products_csv(fp)
                pw.set(100, "Done")
                time.sleep(0.2)
                pw.close()
                self.reload_products()
                messagebox.showinfo("Imported", f"Imported {len(products)} products.")
            threading.Thread(target=task, daemon=True).start()
        except Exception as e:
            messagebox.showerror("Import failed", str(e))
            logging.exception("Import failed")

    def export_products_ui(self):
        fp = filedialog.asksaveasfilename(defaultextension=".csv")
        if not fp:
            return
        try:
            export_products_csv(fp, self.products)
            messagebox.showinfo("Exported", f"Products exported to {fp}")
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            logging.exception("Export failed")

    # ---------------- Products Master Tab ----------------
    def build_products_tab(self):
        # Left pane: search + list
        left = tk.Frame(self.products_frame, width=420, bg=self.current_theme["panel"])
        left.pack(side="left", fill="y", padx=8, pady=8)
        # Search
        tk.Label(left, text="Search:").pack(anchor="w", padx=6, pady=(6,0))
        self.prod_search_var = tk.StringVar()
        sv = tk.Entry(left, textvariable=self.prod_search_var, width=40)
        sv.pack(padx=6)
        sv.bind("<KeyRelease>", lambda e: self.refresh_products_view())

        # Filters frame
        fframe = tk.Frame(left)
        fframe.pack(fill="x", padx=6, pady=6)
        tk.Label(fframe, text="Category").grid(row=0, column=0, sticky="w")
        self.prod_cat_var = tk.StringVar()
        self.prod_cat_cb = ttk.Combobox(fframe, textvariable=self.prod_cat_var, values=self._get_categories_list(), state="readonly", width=20)
        self.prod_cat_cb.grid(row=0, column=1, padx=6, pady=2)
        tk.Label(fframe, text="Brand").grid(row=1, column=0, sticky="w")
        self.prod_brand_var = tk.StringVar()
        tk.Entry(fframe, textvariable=self.prod_brand_var, width=22).grid(row=1, column=1, padx=6, pady=2)
        tk.Button(left, text="Clear Filters", command=self._clear_product_filters).pack(padx=6, pady=(0,6))

        # Products list (tree)
        cols = ("Name","Category","Brand","Rate","Qty")
        self.prod_tree = ttk.Treeview(left, columns=cols, show="headings", height=20)
        for c in cols:
            self.prod_tree.heading(c, text=c)
            self.prod_tree.column(c, width=80 if c!="Name" else 200, anchor="w")
        self.prod_tree.pack(fill="both", expand=True, padx=6, pady=6)
        self.prod_tree.bind("<<TreeviewSelect>>", lambda e: self.on_product_select())
        self.prod_tree.bind("<Double-1>", lambda e: self.open_edit_product_dialog())

        # Buttons under list
        btnf = tk.Frame(left)
        btnf.pack(fill="x", padx=6, pady=6)
        tk.Button(btnf, text="Add Product", command=self.open_add_product_dialog).pack(side="left", padx=4)
        tk.Button(btnf, text="Edit", command=self.open_edit_product_dialog).pack(side="left", padx=4)
        tk.Button(btnf, text="Delete", command=self.delete_selected_product).pack(side="left", padx=4)
        tk.Button(btnf, text="Import CSV", command=self.import_products_ui).pack(side="left", padx=4)
        tk.Button(btnf, text="Export CSV", command=self.export_products_ui).pack(side="left", padx=4)

        # Right pane: details + image + stock management
        right = tk.Frame(self.products_frame, bg=self.current_theme["bg"])
        right.pack(side="left", fill="both", expand=True, padx=8, pady=8)

        # Product details frame
        details = tk.LabelFrame(right, text="Product Details", padx=8, pady=8)
        details.pack(fill="x", padx=6, pady=6)
        # fields
        labels = ["SKU","Name","Category","Brand","Size","Color","HSN","MRP","Rate","Wholesale","SuperWholesale","Discount%","Qty"]
        self.prod_fields = {}
        for i,lbl in enumerate(labels):
            r = i//2
            c = (i%2)*2
            tk.Label(details, text=lbl+(":")).grid(row=r, column=c, sticky="w", padx=4, pady=3)
            ent = tk.Entry(details, width=28)
            ent.grid(row=r, column=c+1, sticky="w", padx=4, pady=3)
            self.prod_fields[lbl.lower()] = ent

        # image panel
        img_frame = tk.LabelFrame(right, text="Image", padx=8, pady=8)
        img_frame.pack(fill="x", padx=6, pady=6)
        self.prod_image_label = tk.Label(img_frame, text="No image", width=40, height=8, bg="#fff")
        self.prod_image_label.pack(side="left", padx=6)
        img_btns = tk.Frame(img_frame)
        img_btns.pack(side="left", padx=6)
        tk.Button(img_btns, text="Attach Image", command=self.attach_image_to_selected).pack(fill="x", pady=4)
        tk.Button(img_btns, text="Clear Image", command=self.clear_image_from_selected).pack(fill="x", pady=4)

        # stock actions
        stock_frame = tk.LabelFrame(right, text="Stock / Actions", padx=8, pady=8)
        stock_frame.pack(fill="x", padx=6, pady=6)
        tk.Button(stock_frame, text="Set Quantity", command=self.set_qty_for_selected).pack(side="left", padx=6)
        tk.Button(stock_frame, text="Adjust Qty (+/-)", command=self.adjust_qty_for_selected).pack(side="left", padx=6)
        tk.Button(stock_frame, text="Bulk Edit (CSV)", command=self.bulk_edit_products_csv).pack(side="left", padx=6)
        tk.Button(stock_frame, text="Backup Products", command=lambda: messagebox.showinfo("Backup", backup_project())).pack(side="left", padx=6)

        # reports (simple)
        rep_frame = tk.LabelFrame(right, text="Reports", padx=8, pady=8)
        rep_frame.pack(fill="x", padx=6, pady=6)
        tk.Button(rep_frame, text="Export sales by date", command=lambda: export_sales_report_csv()).pack(side="left", padx=6)

        # load products into view
        self.refresh_products_view()

    def _get_categories_list(self):
        try:
            conn = sqlite3.connect(DB_FILE)
            cur = conn.cursor()
            cur.execute("SELECT name FROM categories ORDER BY name")
            rows = cur.fetchall()
            conn.close()
            return [r[0] for r in rows]
        except:
            return []

    def _clear_product_filters(self):
        self.prod_search_var.set("")
        self.prod_cat_var.set("")
        self.prod_brand_var.set("")
        self.refresh_products_view()

    def refresh_products_view(self):
        # repopulate categories combobox
        try:
            self.prod_cat_cb['values'] = self._get_categories_list()
        except:
            pass
        q = self.prod_search_var.get().strip().lower() if hasattr(self, "prod_search_var") else ""
        cat = (self.prod_cat_var.get().strip().lower() if hasattr(self, "prod_cat_var") else "")
        brand = (self.prod_brand_var.get().strip().lower() if hasattr(self, "prod_brand_var") else "")
        # filter products
        allp = self.products or {}
        filtered = {}
        for name,p in allp.items():
            name_l = (p.get("name") or name).lower()
            if q and q not in name_l and q not in (p.get("brand","") or "").lower() and q not in (p.get("category","") or "").lower():
                continue
            if cat and cat != "" and cat not in ((p.get("category") or "").lower()):
                continue
            if brand and brand != "" and brand not in ((p.get("brand") or "").lower()):
                continue
            filtered[name] = p
        # clear tree
        for i in self.prod_tree.get_children():
            self.prod_tree.delete(i)
        # insert
        for name,p in sorted(filtered.items(), key=lambda x: x[0])[:10000]:
            self.prod_tree.insert("", tk.END, iid=name, values=(p.get("name") or name, p.get("category") or "", p.get("brand") or "", p.get("rate") or 0, p.get("qty") or 0))

    def on_product_select(self):
        sel = self.prod_tree.selection()
        if not sel:
            return
        key = sel[0]
        p = self.products.get(key.lower()) or self.products.get(key)
        if not p:
            return
        # populate detail fields
        try:
            for k,ent in self.prod_fields.items():
                ent.delete(0, tk.END)
                ent.insert(0, str(p.get(k) if p.get(k) is not None else ""))
        except:
            pass
        # image preview
        img = p.get("image_path")
        if img and os.path.exists(img) and Image and ImageTk:
            try:
                im = Image.open(img)
                im.thumbnail((200,200))
                self._prod_img_tk = ImageTk.PhotoImage(im)
                self.prod_image_label.configure(image=self._prod_img_tk, text="")
            except:
                self.prod_image_label.configure(image="", text="Image load failed")
        else:
            self.prod_image_label.configure(image="", text="No image")

    def open_add_product_dialog(self):
        # simple dialog form
        dlg = tk.Toplevel(self.root)
        dlg.title("Add Product")
        fields = ["sku","name","category","brand","size","color","hsn","mrp","rate","wholesale","super_wholesale","discount","qty","image_path","notes"]
        entries = {}
        for i, f in enumerate(fields):
            tk.Label(dlg, text=f.capitalize()+":").grid(row=i, column=0, sticky="w", padx=6, pady=2)
            e = tk.Entry(dlg, width=40)
            e.grid(row=i, column=1, padx=6, pady=2)
            entries[f] = e
        def on_add():
            p = {k: (entries[k].get().strip() if entries[k].get().strip()!="" else None) for k in fields}
            # convert numeric fields
            for numf in ("mrp","rate","wholesale","super_wholesale","discount"):
                try:
                    p[numf] = float(p.get(numf) or 0)
                except:
                    p[numf] = 0.0
            try:
                p["qty"] = int(float(p.get("qty") or 0))
            except:
                p["qty"] = 0
            if not p.get("name"):
                messagebox.showerror("Error","Name required")
                return
            try:
                save_product_to_db(p)
                self.products[p["name"].lower()] = p
                self.refresh_products_view()
                dlg.destroy()
                messagebox.showinfo("Added","Product added")
            except Exception as e:
                messagebox.showerror("Error", str(e))
        tk.Button(dlg, text="Add", command=on_add).grid(row=len(fields), column=0, columnspan=2, pady=8)

    def open_edit_product_dialog(self):
        sel = self.prod_tree.selection()
        if not sel:
            messagebox.showinfo("Edit", "Select a product then press Edit.")
            return
        key = sel[0]
        p = self.products.get(key.lower()) or self.products.get(key)
        if not p:
            messagebox.showerror("Error","Product not found")
            return
        dlg = tk.Toplevel(self.root)
        dlg.title("Edit Product")
        fields = ["sku","name","category","brand","size","color","hsn","mrp","rate","wholesale","super_wholesale","discount","qty","image_path","notes"]
        entries = {}
        for i, f in enumerate(fields):
            tk.Label(dlg, text=f.capitalize()+":").grid(row=i, column=0, sticky="w", padx=6, pady=2)
            e = tk.Entry(dlg, width=40)
            e.grid(row=i, column=1, padx=6, pady=2)
            e.insert(0, str(p.get(f) or ""))
            entries[f] = e
        def on_save():
            newp = {k: (entries[k].get().strip() if entries[k].get().strip()!="" else None) for k in fields}
            for numf in ("mrp","rate","wholesale","super_wholesale","discount"):
                try:
                    newp[numf] = float(newp.get(numf) or 0)
                except:
                    newp[numf] = 0.0
            try:
                newp["qty"] = int(float(newp.get("qty") or 0))
            except:
                newp["qty"] = 0
            if not newp.get("name"):
                messagebox.showerror("Error","Name required")
                return
            try:
                save_product_to_db(newp)
                # update in-memory
                self.products[newp["name"].lower()] = newp
                # if name changed remove old key
                if newp["name"].lower() != key.lower() and key.lower() in self.products:
                    try: del self.products[key.lower()]
                    except: pass
                self.refresh_products_view()
                dlg.destroy()
                messagebox.showinfo("Saved","Product updated")
            except Exception as e:
                messagebox.showerror("Error", str(e))
        tk.Button(dlg, text="Save", command=on_save).grid(row=len(fields), column=0, columnspan=2, pady=8)

    def delete_selected_product(self):
        sel = self.prod_tree.selection()
        if not sel:
            return
        key = sel[0]
        p = self.products.get(key.lower()) or self.products.get(key)
        if not p:
            return
        if not messagebox.askyesno("Confirm", f"Delete product '{p.get('name')}'?"):
            return
        try:
            delete_product_from_db(p.get("name"))
        except:
            pass
        try:
            del self.products[p.get("name").lower()]
        except:
            pass
        self.refresh_products_view()
        messagebox.showinfo("Deleted","Product deleted")

    def attach_image_to_selected(self):
        sel = self.prod_tree.selection()
        if not sel:
            messagebox.showinfo("Attach Image", "Select a product first")
            return
        key = sel[0]
        p = self.products.get(key.lower()) or self.products.get(key)
        if not p:
            return
        fp = filedialog.askopenfilename(filetypes=[("Image files","*.png;*.jpg;*.jpeg;*.gif;*.bmp"),("All files","*.*")])
        if not fp:
            return
        # copy image to local images folder
        img_dir = "product_images"
        os.makedirs(img_dir, exist_ok=True)
        dest = os.path.join(img_dir, os.path.basename(fp))
        try:
            shutil.copyfile(fp, dest)
            p["image_path"] = dest
            save_product_to_db(p)
            self.on_product_select()
            messagebox.showinfo("Image", "Image attached")
        except Exception as e:
            messagebox.showerror("Failed", str(e))

    def clear_image_from_selected(self):
        sel = self.prod_tree.selection()
        if not sel:
            return
        key = sel[0]
        p = self.products.get(key.lower()) or self.products.get(key)
        if not p:
            return
        p["image_path"] = None
        save_product_to_db(p)
        self.on_product_select()
        messagebox.showinfo("Image", "Image cleared")

    def set_qty_for_selected(self):
        sel = self.prod_tree.selection()
        if not sel:
            return
        key = sel[0]
        p = self.products.get(key.lower()) or self.products.get(key)
        if not p:
            return
        try:
            newq = simpledialog.askinteger("Set Quantity", "Enter new quantity:", initialvalue=p.get("qty",0), parent=self.root)
            if newq is None:
                return
            p["qty"] = int(newq)
            save_product_to_db(p)
            self.refresh_products_view()
            messagebox.showinfo("Qty updated", f"Qty set to {p['qty']}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def adjust_qty_for_selected(self):
        sel = self.prod_tree.selection()
        if not sel:
            return
        key = sel[0]
        p = self.products.get(key.lower()) or self.products.get(key)
        if not p:
            return
        try:
            delta = simpledialog.askinteger("Adjust Qty", "Enter quantity delta (+/-):", initialvalue=0, parent=self.root)
            if delta is None:
                return
            p["qty"] = int((p.get("qty") or 0) + int(delta))
            save_product_to_db(p)
            self.refresh_products_view()
            messagebox.showinfo("Qty updated", f"Qty is now {p['qty']}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def bulk_edit_products_csv(self):
        fp = filedialog.askopenfilename(filetypes=[("CSV Files","*.csv")])
        if not fp:
            return
        try:
            count = 0
            with open(fp, newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for r in reader:
                    name = r.get("name") or r.get("Name")
                    if not name:
                        continue
                    p = load_products_from_db().get(name.lower(), {})
                    # update fields present
                    for k in ("mrp","rate","wholesale","super_wholesale","discount","qty","category","brand"):
                        if k in r and r[k]!="":
                            if k=="qty":
                                try: p["qty"]=int(float(r[k]))
                                except: pass
                            else:
                                try: p[k]=float(r[k]) if k in ("mrp","rate","wholesale","super_wholesale","discount") else r[k]
                                except: p[k]=r[k]
                    p["name"]=name
                    save_product_to_db(p)
                    count += 1
            self.reload_products()
            messagebox.showinfo("Bulk edit", f"Processed {count} rows")
        except Exception as e:
            messagebox.showerror("Bulk edit failed", str(e))
            logging.exception("Bulk edit failed")

    # ---------------- Product tab helper ----------------
    def open_products_tab(self):
        self.notebook.select(self.products_frame)

    # ---------------- Utility functions (other parts) ----------------
    def show_shortcuts_help(self):
        shortcuts = [
            "Ctrl+N — New client / clear invoice",
            "Ctrl+P — Open Print tab & refresh preview",
            "Ctrl+Shift+P — Open Products Master",
            "Ctrl+S — Save bill as PDF",
            "Ctrl+Shift+S — Save customer",
            "Delete — Delete selected item",
            "Ctrl+D — Duplicate selected item",
            "Ctrl+E or Double-click — Edit selected item",
            "Ctrl+Up / Ctrl+Down — Move selected item up / down",
            "F5 — Refresh printer list",
            "F9 — Reload product master",
            "Alt+T — Cycle theme",
            "Ctrl+L — Focus LowStockThreshold entry",
            "Ctrl+I — Focus Item name (start adding)",
            "Ctrl+B — Focus Bill preview (print tab)",
            "Ctrl+R — Refresh totals",
            "Ctrl+M — Merge PDFs (opens file dialog)",
            "Ctrl+H — Show this help"
        ]
        messagebox.showinfo("Keyboard Shortcuts", "\n".join(shortcuts))

    # ---------------- Import/Export adapters ----------------
    def import_products_ui(self):
        fp = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")])
        if not fp:
            return
        try:
            import_products_csv(fp)
            self.products = load_products_from_db()
            self.reload_products()
            messagebox.showinfo("Imported", "Products imported successfully.")
        except Exception as e:
            messagebox.showerror("Import failed", str(e))
            logging.exception("Import failed")

    def export_products_ui(self):
        fp = filedialog.asksaveasfilename(defaultextension=".csv")
        if not fp:
            return
        try:
            export_products_csv(fp, self.products)
            messagebox.showinfo("Exported", f"Products exported to {fp}")
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            logging.exception("Export failed")

# ------------------- Additional DB bill save -------------------
def save_bill_to_db(app, db_file=DB_FILE):
    conn = sqlite3.connect(db_file)
    cur = conn.cursor()
    bill_no, _ = get_next_bill_no()
    cur.execute("INSERT INTO bills (bill_no, date, customer, phone, subtotal, gst, total, paid, due) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (bill_no, datetime.now().isoformat(), app.cust_name.get(), app.cust_phone.get(), app.sub_total,
                 round(app.sub_total * app.gst_percent.get() / 100, 2), app.total, app.paid_amount.get(), app.total - app.paid_amount.get()))
    bill_id = cur.lastrowid
    for it in app.items_in_bill:
        cur.execute("INSERT INTO bill_items (bill_id, product_name, mrp, rate, discount, qty, total) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (bill_id, it["name"], it["mrp"], it["rate"], it["discount"], it["qty"], it["total"]))
    conn.commit()
    conn.close()
    return bill_id

# ------------------- Reports -------------------
def sales_report_by_date(db_file=DB_FILE):
    conn = sqlite3.connect(db_file)
    cur = conn.cursor()
    cur.execute("""
    SELECT substr(date,1,10) as day, COUNT(*) as bills, SUM(total) as total_amount
    FROM bills
    GROUP BY day
    ORDER BY day DESC
    """)
    rows = cur.fetchall()
    conn.close()
    return rows

def export_sales_report_csv(out_file="sales_by_date.csv", db_file=DB_FILE):
    rows = sales_report_by_date(db_file=db_file)
    with open(out_file, "w", newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(["date", "num_bills", "total_amount"])
        for r in rows:
            w.writerow(r)
    return out_file

# ------------------- Users & Auth (tiny) -------------------
DEFAULT_USERS = {"admin": {"password": "admin", "role": "admin"}, "cashier": {"password": "cashier", "role": "cashier"}}
def load_users():
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(DEFAULT_USERS, f, indent=2)
    return DEFAULT_USERS.copy()
def authenticate(username, password):
    users = load_users()
    u = users.get(username)
    if not u:
        return False, None
    if u.get("password") == password:
        return True, u.get("role")
    return False, None

# ------------------- Preferences UI -------------------
def open_preferences_dialog(parent, settings):
    win = tk.Toplevel(parent)
    win.title("Preferences")
    win.geometry("420x240")
    tk.Label(win, text="Auto Backup:").grid(row=0, column=0, sticky="w", padx=8, pady=8)
    ab_var = tk.BooleanVar(value=settings.get("auto_backup", False))
    tk.Checkbutton(win, variable=ab_var, text="Enable automatic backups").grid(row=0, column=1, sticky="w")
    tk.Label(win, text="Interval (minutes):").grid(row=1, column=0, sticky="w", padx=8, pady=8)
    interval_var = tk.IntVar(value=settings.get("auto_backup_interval_minutes", 60))
    tk.Entry(win, textvariable=interval_var, width=8).grid(row=1, column=1, sticky="w")
    tk.Label(win, text="Page size (pagination):").grid(row=2, column=0, sticky="w", padx=8, pady=8)
    page_var = tk.IntVar(value=settings.get("page_size", 200))
    tk.Entry(win, textvariable=page_var, width=8).grid(row=2, column=1, sticky="w")
    def on_save():
        settings["auto_backup"] = ab_var.get()
        settings["auto_backup_interval_minutes"] = int(interval_var.get())
        settings["page_size"] = int(page_var.get())
        save_settings(settings)
        messagebox.showinfo("Saved", "Preferences saved.")
        win.destroy()
    ttk.Button(win, text="Save", command=on_save).grid(row=6, column=0, columnspan=2, pady=12)

# ------------------- Run App -------------------
if __name__ == "__main__":
    try:
        init_db()
    except Exception:
        pass
    settings = load_settings()
    root = tk.Tk()
    app = BillingApp(root)
    try:
        app.apply_theme(settings.get("last_theme", app.current_theme_name))
    except:
        app.apply_theme(app.current_theme_name)
    root.mainloop()
